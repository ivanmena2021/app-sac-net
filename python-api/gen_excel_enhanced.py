"""
Enhanced Excel export with professional formatting for SAC agricultural insurance reports.
Replaces the inline _build_consolidated_excel that just does df.to_excel().
"""

import io
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ── Color palette ──────────────────────────────────────────────────────────
NAVY = "0C2340"
TEAL = "1A5276"
CREAM = "F8F9FA"
WHITE = "FFFFFF"
GREEN_LIGHT = "D5F5E3"
GREEN_DARK = "27AE60"

# ── Reusable styles ───────────────────────────────────────────────────────
_thin_side = Side(style="thin", color="CCCCCC")
BORDER_THIN = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

ACCENT_FILL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=CREAM, end_color=CREAM, fill_type="solid")
GREEN_FILL = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")

TITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=14)
KPI_LABEL_FONT = Font(name="Calibri", bold=True, color=TEAL, size=10)
KPI_VALUE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=13)
BODY_FONT = Font(name="Calibri", size=10)


def _auto_column_width(ws, min_width=10, max_width=35):
    """Set column widths based on content length."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _apply_header_style(ws, row, col_start, col_end):
    """Apply navy header style to a row range."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN


def _write_dataframe(ws, df, start_row=1, money_cols=None, int_cols=None, decimal_cols=None):
    """Write a DataFrame to a worksheet with formatting. Uses batch column writes."""
    money_cols = money_cols or []
    int_cols = int_cols or []
    decimal_cols = decimal_cols or []

    # Write headers
    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN

    # Write data by column (batch approach)
    for c_idx, col_name in enumerate(df.columns, 1):
        values = df[col_name].values
        is_money = col_name in money_cols
        is_int = col_name in int_cols
        is_decimal = col_name in decimal_cols

        for r_offset, val in enumerate(values):
            r = start_row + 1 + r_offset
            # Handle numpy/pandas types
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None
            elif pd.isna(val):
                val = None

            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left",
                                       vertical="center")

            # Number formats
            if is_money and val is not None:
                cell.number_format = '#,##0.00'
            elif is_int and val is not None:
                cell.number_format = '#,##0'
            elif is_decimal and val is not None:
                cell.number_format = '#,##0.00'

            # Alternating row color
            if r_offset % 2 == 1:
                cell.fill = ALT_ROW_FILL

    return start_row + len(df)


def _build_resumen_sheet(wb, datos):
    """Build the 'Resumen' summary sheet."""
    ws = wb.active
    ws.title = "Resumen"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws.cell(row=1, column=1, value="Seguro Agricola Catastrofico \u2014 SAC 2025-2026")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle with fecha_corte
    fecha_corte = datos.get("fecha_corte", "")
    ws.merge_cells("A2:H2")
    sub_cell = ws.cell(row=2, column=1, value=f"Fecha de corte: {fecha_corte}")
    sub_cell.font = Font(name="Calibri", italic=True, color=TEAL, size=10)
    sub_cell.alignment = Alignment(horizontal="center")

    # KPI section — row 3 labels, row 4 values
    kpis = [
        ("Total Avisos", datos.get("total_avisos", 0), "#,##0"),
        ("Ha Indemnizadas", datos.get("ha_indemnizadas", 0), "#,##0.00"),
        ("Monto Indemnizado (S/)", datos.get("monto_indemnizado", 0), '#,##0.00'),
        ("Siniestralidad (%)", datos.get("indice_siniestralidad", 0), "0.0%"),
    ]

    for i, (label, value, fmt) in enumerate(kpis):
        col_start = 1 + i * 2
        col_end = col_start + 1

        ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
        lbl_cell = ws.cell(row=3, column=col_start, value=label)
        lbl_cell.font = KPI_LABEL_FONT
        lbl_cell.alignment = Alignment(horizontal="center")
        lbl_cell.fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type="solid")

        ws.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_end)
        val_cell = ws.cell(row=4, column=col_start, value=value)
        val_cell.font = KPI_VALUE_FONT
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.number_format = fmt

    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 30

    # Additional KPIs row 5-6
    extra_kpis = [
        ("Monto Desembolsado (S/)", datos.get("monto_desembolsado", 0), '#,##0.00'),
        ("Productores Desembolso", datos.get("productores_desembolso", 0), "#,##0"),
        ("% Desembolso", datos.get("pct_desembolso", 0), "0.0%"),
        ("Prima Total (S/)", datos.get("prima_total", 0), '#,##0.00'),
    ]

    for i, (label, value, fmt) in enumerate(extra_kpis):
        col_start = 1 + i * 2
        col_end = col_start + 1

        ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_end)
        lbl_cell = ws.cell(row=5, column=col_start, value=label)
        lbl_cell.font = KPI_LABEL_FONT
        lbl_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=6, start_column=col_start, end_row=6, end_column=col_end)
        val_cell = ws.cell(row=6, column=col_start, value=value)
        val_cell.font = KPI_VALUE_FONT
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.number_format = fmt

    # Department ranking table from cuadro2
    cuadro2 = datos.get("cuadro2")
    if cuadro2 is not None and not cuadro2.empty:
        money_cols = [c for c in cuadro2.columns if any(k in c.upper() for k in
                      ["MONTO", "INDEMNIZA", "PRIMA", "DESEMBOL"])]
        int_cols = [c for c in cuadro2.columns if any(k in c.upper() for k in
                    ["AVISO", "PRODUCTOR", "TOTAL"])]
        decimal_cols = [c for c in cuadro2.columns if any(k in c.upper() for k in
                        ["HA", "HECTA", "SINIEST", "INDICE"])]
        _write_dataframe(ws, cuadro2, start_row=8,
                         money_cols=money_cols, int_cols=int_cols, decimal_cols=decimal_cols)

    _auto_column_width(ws)


def _build_data_sheet(wb, df, sheet_name):
    """Build a data sheet (Consolidado or per-company) with full formatting."""
    ws = wb.create_sheet(title=sheet_name)

    if df is None or df.empty:
        ws.cell(row=1, column=1, value="Sin datos disponibles")
        return

    # Determine column types by name heuristics
    money_cols = [c for c in df.columns if any(k in c.upper() for k in
                  ["MONTO", "INDEMNIZA", "PRIMA", "DESEMBOL", "COSTO", "VALOR", "PAGO"])]
    int_cols = [c for c in df.columns if any(k in c.upper() for k in
                ["AVISO", "PRODUCTOR", "TOTAL", "NRO", "NUMERO", "CANTIDAD", "POLIZA"])]
    decimal_cols = [c for c in df.columns if any(k in c.upper() for k in
                    ["HA_", "HECTA", "SINIEST", "INDICE", "PORCENTAJE", "PCT"])]

    last_row = _write_dataframe(ws, df, start_row=1,
                                money_cols=money_cols, int_cols=int_cols,
                                decimal_cols=decimal_cols)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    last_col = get_column_letter(len(df.columns))
    ws.auto_filter.ref = f"A1:{last_col}{last_row + 1}"

    # Conditional formatting: green fill for INDEMNIZACION > 0
    indem_cols = [i + 1 for i, c in enumerate(df.columns)
                  if "INDEMNIZA" in c.upper()]
    for c_idx in indem_cols:
        col_letter = get_column_letter(c_idx)
        for r in range(2, last_row + 2):
            cell = ws.cell(row=r, column=c_idx)
            if cell.value is not None and isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = GREEN_FILL

    _auto_column_width(ws)


def generate_enhanced_excel(datos):
    """Generate a professional Excel workbook with multiple sheets.

    Parameters
    ----------
    datos : dict
        Keys used:
        - midagri: DataFrame with all claims data
        - total_avisos, ha_indemnizadas, monto_indemnizado, monto_desembolsado,
          productores_desembolso, indice_siniestralidad, pct_desembolso, prima_total,
          fecha_corte
        - cuadro2: DataFrame with department indemnification summary

    Returns
    -------
    bytes
        XLSX file content.
    """
    wb = Workbook()

    # Sheet 1: Resumen
    _build_resumen_sheet(wb, datos)

    # Sheet 2: Consolidado
    midagri = datos.get("midagri")
    if midagri is not None:
        _build_data_sheet(wb, midagri, "Consolidado")
    else:
        ws = wb.create_sheet(title="Consolidado")
        ws.cell(row=1, column=1, value="Sin datos disponibles")

    # Sheet 3 & 4: Per-company sheets
    empresa_col = None
    if midagri is not None:
        for col in midagri.columns:
            if "EMPRESA" in col.upper():
                empresa_col = col
                break

    if empresa_col and midagri is not None:
        # La Positiva
        mask_lp = midagri[empresa_col].str.upper().str.contains("POSITIVA", na=False)
        df_lp = midagri[mask_lp].reset_index(drop=True)
        _build_data_sheet(wb, df_lp, "La Positiva")

        # Rimac
        mask_ri = midagri[empresa_col].str.upper().str.contains("RIMAC", na=False)
        df_ri = midagri[mask_ri].reset_index(drop=True)
        _build_data_sheet(wb, df_ri, "Rimac")
    else:
        wb.create_sheet(title="La Positiva")
        wb.create_sheet(title="Rimac")

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

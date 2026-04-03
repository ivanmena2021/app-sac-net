"""
gen_excel_eme.py — Generador del Formato Reporte EME (Excel).
Replica la estructura del archivo formato_reporte_EME.
"""

import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _fmt_num(val, decimals=2):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return 0
    try:
        return round(float(val), decimals)
    except (ValueError, TypeError):
        return 0


def generate_reporte_eme(datos):
    """
    Genera el Excel del reporte EME (formato_reporte_EME).
    Retorna bytes del archivo .xlsx.
    """
    midagri = datos["midagri"]
    materia = datos["materia"]

    # Agrupar por departamento (= REGIÓN)
    dept_data = midagri.groupby("DEPARTAMENTO").agg(
        n_avisos=("DEPARTAMENTO", "count"),
        sup_indemnizada=("SUP_INDEMNIZADA", "sum"),
        indemnizacion=("INDEMNIZACION", "sum"),
        monto_desembolsado=("MONTO_DESEMBOLSADO", "sum"),
        n_productores=("N_PRODUCTORES", "sum"),
    ).reset_index()

    # Distritos por departamento
    if "DISTRITO" in midagri.columns and "PROVINCIA" in midagri.columns:
        dist_info = midagri.groupby("DEPARTAMENTO").apply(
            lambda g: _build_district_text(g)
        ).reset_index()
        dist_info.columns = ["DEPARTAMENTO", "DISTRITOS"]
        dept_data = dept_data.merge(dist_info, on="DEPARTAMENTO", how="left")
    else:
        dept_data["DISTRITOS"] = ""

    # Empresa aseguradora y prima total
    if "EMPRESA_ASEGURADORA" in materia.columns:
        emp_map = materia.set_index("DEPARTAMENTO")[["EMPRESA_ASEGURADORA", "PRIMA_TOTAL"]].to_dict("index")
    else:
        emp_map = {}

    # Tipo siniestro predominante por departamento
    if "TIPO_SINIESTRO" in midagri.columns:
        tipo_dict = {}
        for depto_name, group in midagri.groupby("DEPARTAMENTO"):
            tipo_dict[depto_name] = group["TIPO_SINIESTRO"].value_counts().head(3).to_dict()
        tipo_info = pd.DataFrame([
            {"DEPARTAMENTO": k, "TOP_SINIESTROS": v} for k, v in tipo_dict.items()
        ])
    else:
        tipo_info = pd.DataFrame(columns=["DEPARTAMENTO", "TOP_SINIESTROS"])

    dept_data = dept_data.merge(tipo_info, on="DEPARTAMENTO", how="left")

    # ═══ Crear Workbook ═══
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Estilos
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_font = Font(name="Calibri", size=9)
    number_format_soles = '#,##0.00'
    number_format_int = '#,##0'
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "REGIÓN",
        "DISTRITOS",
        "BENEFICIARIOS",
        "ACCIÓN IMPLEMENTADA\n- MONTO DESEMBOLSADO",
        "ACCION EN IMPLEMENTACIÓN\n - MONTO INDEMNIZADO ",
        "ACCION POR IMPLEMENTAR - PRIMA TOTAL",
        "DESCRIPCIÓN",
        "UNIDAD RESPONSABLE",
        "CUANTIFICACIÓN/ TOTAL - HAS INDEMNIZADAS",
        "OBSERVACIONES",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    dept_data = dept_data.sort_values("DEPARTAMENTO")

    for row_idx, (_, row) in enumerate(dept_data.iterrows(), 2):
        depto = row["DEPARTAMENTO"]
        depto_upper = depto.strip().upper()

        # Datos estáticos
        emp_info = emp_map.get(depto_upper, {})
        empresa = emp_info.get("EMPRESA_ASEGURADORA", "N/D")
        prima_total = _fmt_num(emp_info.get("PRIMA_TOTAL", 0))

        # Descripción
        top_sin = row.get("TOP_SINIESTROS", {})
        if isinstance(top_sin, dict) and top_sin:
            tipos_text = ", ".join([f"{k.lower()} ({v})" for k, v in top_sin.items()])
        else:
            tipos_text = "sin datos"

        descripcion = (
            f"Aseguradora: {empresa}. {int(row['n_avisos'])} avisos de siniestro. "
            f"Principales siniestros: {tipos_text}."
        )

        # Observaciones
        if row["monto_desembolsado"] > 0:
            obs = f"Desembolsos realizados a {int(row['n_productores'])} productores."
        elif row["indemnizacion"] > 0:
            obs = "Indemnizaciones reconocidas, pendiente de desembolso."
        else:
            obs = "En proceso de evaluación."

        # Escribir fila
        values = [
            depto.title(),
            row.get("DISTRITOS", ""),
            int(row["n_productores"]) if row["n_productores"] > 0 else 0,
            _fmt_num(row["monto_desembolsado"]),
            _fmt_num(row["indemnizacion"]),
            prima_total,
            descripcion,
            "DGASFS / DSFFA",
            _fmt_num(row["sup_indemnizada"]),
            obs,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Formato numérico
            if col_idx in [3]:
                cell.number_format = number_format_int
                cell.alignment = Alignment(horizontal="right", vertical="top")
            elif col_idx in [4, 5, 6, 9]:
                cell.number_format = number_format_soles
                cell.alignment = Alignment(horizontal="right", vertical="top")

    # Anchos de columna
    col_widths = [15, 50, 13, 18, 18, 18, 50, 18, 18, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Altura de fila header
    ws.row_dimensions[1].height = 40

    # Guardar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_district_text(group):
    """Construye texto resumen de distritos por provincia."""
    try:
        prov_dist = group.groupby("PROVINCIA")["DISTRITO"].nunique()
        parts = []
        for prov, n in prov_dist.sort_values(ascending=False).items():
            parts.append(f"{str(prov).title()} ({n})")
        n_dist = group["DISTRITO"].nunique()
        n_prov = group["PROVINCIA"].nunique()
        return f"{n_dist} distritos en {n_prov} provincias: {', '.join(parts)}"
    except Exception:
        return ""

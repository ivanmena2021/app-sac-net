"""
data_processor.py — Procesamiento de datos dinámicos (MIDAGRI y Siniestros)
combinados con datos estáticos (Materia Asegurada y Resumen SAC).
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime

STATIC_DIR = os.path.join(os.path.dirname(__file__), "static_data")

# ─── Tipos de siniestro asociados a lluvias intensas ───
LLUVIA_TYPES = {"INUNDACION", "INUNDACIÓN", "HUAYCO", "LLUVIAS EXCESIVAS", "DESLIZAMIENTO", "DESLIZAMIENTOS"}

# ─── Mapeo capital → departamento (la columna A del Materia Asegurada viene vacía) ───
CAPITAL_TO_DEPTO = {
    "CHACHAPOYAS": "AMAZONAS",
    "HUARAZ": "ANCASH",
    "AREQUIPA": "AREQUIPA",
    "CUSCO": "CUSCO",
    "HUANCAVELICA": "HUANCAVELICA",
    "HUÁNUCO": "HUANUCO",
    "HUANUCO": "HUANUCO",
    "HUANCAYO": "JUNIN",
    "TRUJILLO": "LA LIBERTAD",
    "CHICLAYO": "LAMBAYEQUE",
    "LIMA": "LIMA",
    "IQUITOS": "LORETO",
    "MADRE DE DIOS": "MADRE DE DIOS",
    "PUERTO MALDONADO": "MADRE DE DIOS",
    "CERRO DE PASCO": "PASCO",
    "PIURA": "PIURA",
    "PUNO": "PUNO",
    "MOYOBAMBA": "SAN MARTIN",
    "TACNA": "TACNA",
    "PUCALLPA": "UCAYALI",
    "ABANCAY": "APURIMAC",
    "AYACUCHO": "AYACUCHO",
    "CAJAMARCA": "CAJAMARCA",
    "ICA": "ICA",
    "MOQUEGUA": "MOQUEGUA",
    "TUMBES": "TUMBES",
}


def load_primas_historicas():
    """Carga primas netas por departamento y campaña desde Excel estático.

    Fuente: Primas_Totales_SAC_2020-2026.xlsx (6 hojas, una por campaña).
    Retorna: {campaña: {depto_upper: prima_neta_float}}
    Ver METODOLOGIA_DATOS.md para contexto.
    """
    path = os.path.join(STATIC_DIR, "Primas_Totales_SAC_2020-2026.xlsx")
    if not os.path.exists(path):
        return {}

    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}

    # Mapeo de nombre de hoja a campaña estándar
    sheet_to_camp = {
        "SAC 2020-2021": "2020-2021", "SAC 2021-2022": "2021-2022",
        "SAC 2022-2023": "2022-2023", "SAC 2023-2024": "2023-2024",
        "SAC 2024-2025": "2024-2025", "SAC 2025-2026": "2025-2026",
    }

    for sheet_name, camp in sheet_to_camp.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        camp_data = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            depto_raw = str(row[0]).strip().upper()
            if "TOTAL" in depto_raw or "REGION" in depto_raw or "DEPARTAMENTO" in depto_raw:
                continue

            # Normalizar acentos
            for a, p in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N"),
                         ("á","A"),("é","E"),("í","I"),("ó","O"),("ú","U"),("ñ","N")]:
                depto_raw = depto_raw.replace(a, p)

            # Prima neta es la última columna numérica
            numerics = [v for v in row if isinstance(v, (int, float)) and v > 0]
            if numerics:
                camp_data[depto_raw] = float(numerics[-1])

        result[camp] = camp_data

    wb.close()
    return result


def load_materia_asegurada():
    """Carga Materia Asegurada (datos estáticos de póliza por departamento)."""
    from openpyxl import load_workbook
    path = os.path.join(STATIC_DIR, "Materia_Asegurada_SAC_2025-2026.xlsx")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Encontrar fila de header (contiene "Capital")
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=False), 1):
        vals = [str(c.value).strip() for c in row if c.value is not None]
        if "Capital" in vals or "Departamento" in vals:
            header_row = row_idx
            break
    if header_row is None:
        header_row = 3  # default

    # Leer datos fila por fila
    headers_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(headers_cells)]

    data_rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)
        # Saltar filas completamente vacías
        if all(v is None for v in vals):
            continue
        data_rows.append(vals)

    df = pd.DataFrame(data_rows, columns=headers)

    # Renombrar columnas
    col_map = {}
    for c in df.columns:
        cs = str(c).strip().upper()
        if "DEPARTAMENTO" in cs:
            col_map[c] = "DEPARTAMENTO"
        elif "CAPITAL" in cs:
            col_map[c] = "CAPITAL"
        elif "EMPRESA" in cs:
            col_map[c] = "EMPRESA_ASEGURADORA"
        elif "CULTIVOS" in cs:
            col_map[c] = "CULTIVOS_ASEGURADOS"
        elif "PRIMA TOTAL" in cs:
            col_map[c] = "PRIMA_TOTAL"
        elif "PRIMA NETA" in cs:
            col_map[c] = "PRIMA_NETA"
        elif "SUPERFICIE ASEGURADA" in cs:
            col_map[c] = "SUPERFICIE_ASEGURADA"
        elif "PRODUCTORES" in cs:
            col_map[c] = "PRODUCTORES_ASEGURADOS"
        elif "VALORES" in cs:
            col_map[c] = "VALORES_ASEGURADOS"
        elif "DISPARADOR" in cs:
            col_map[c] = "DISPARADOR"
        elif "SUMA ASEGURADA" in cs:
            col_map[c] = "SUMA_ASEGURADA_HA"
    df = df.rename(columns=col_map)

    # Asignar departamento a partir de la capital (la col DEPARTAMENTO viene vacía)
    if "CAPITAL" in df.columns:
        df["CAPITAL"] = df["CAPITAL"].astype(str).str.strip().str.upper()
        df["DEPARTAMENTO"] = df["CAPITAL"].map(CAPITAL_TO_DEPTO)
        # Remover fila TOTAL y filas sin capital válida
        df = df[df["DEPARTAMENTO"].notna()]

    # Convertir numéricas
    for col in ["PRIMA_TOTAL", "PRIMA_NETA", "SUPERFICIE_ASEGURADA", "PRODUCTORES_ASEGURADOS", "VALORES_ASEGURADOS"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def load_resumen_sac():
    """Carga Resumen SAC (datos estáticos de resumen)."""
    path = os.path.join(STATIC_DIR, "Resumen_SAC_2025-2026.xlsx")
    sheets = {}
    # Hoja Primas y Cobertura
    try:
        df = pd.read_excel(path, header=None, sheet_name="Primas y Cobertura")
        header_row = None
        for i, row in df.iterrows():
            vals = [str(v).strip() for v in row.values if pd.notna(v)]
            if "Departamento" in vals:
                header_row = i
                break
        if header_row is not None:
            df.columns = df.iloc[header_row].values
            df = df.iloc[header_row + 1:].reset_index(drop=True)
            df = df.dropna(subset=[df.columns[0]])
            sheets["primas"] = df
    except Exception:
        pass

    # Hoja Indemnizaciones y Desembolsos
    try:
        df = pd.read_excel(path, header=None, sheet_name="Indemnizaciones y Desembolsos")
        header_row = None
        for i, row in df.iterrows():
            vals = [str(v).strip() for v in row.values if pd.notna(v)]
            if "Departamento" in vals:
                header_row = i
                break
        if header_row is not None:
            df.columns = df.iloc[header_row].values
            df = df.iloc[header_row + 1:].reset_index(drop=True)
            df = df.dropna(subset=[df.columns[0]])
            sheets["indemnizaciones"] = df
    except Exception:
        pass

    return sheets


def _normalize_midagri(uploaded_bytes):
    """Normaliza archivo MIDAGRI (tiene header en fila 1, datos desde fila 2)."""
    df = pd.read_excel(uploaded_bytes, header=None, sheet_name=0)
    # Encontrar fila de header real
    header_row = None
    for i, row in df.iterrows():
        vals = [str(v).strip().upper() for v in row.values if pd.notna(v)]
        if any("CAMPAÑA" in v or "CÓDIGO DE AVISO" in v or "CODIGO DE AVISO" in v for v in vals):
            header_row = i
            break
    if header_row is None:
        header_row = 1  # default

    df.columns = [str(c).strip() for c in df.iloc[header_row].values]
    df = df.iloc[header_row + 1:].reset_index(drop=True)

    # Estandarizar nombres de columnas
    col_map = {}
    for c in df.columns:
        cu = str(c).strip().upper()
        if "CAMPAÑA" in cu:
            col_map[c] = "CAMPAÑA"
        elif "CÓDIGO DE AVISO" in cu or "CODIGO DE AVISO" in cu:
            col_map[c] = "CODIGO_AVISO"
        elif cu == "DEPARTAMENTO":
            col_map[c] = "DEPARTAMENTO"
        elif cu == "PROVINCIA":
            col_map[c] = "PROVINCIA"
        elif cu == "DISTRITO":
            col_map[c] = "DISTRITO"
        elif "SECTOR" in cu:
            col_map[c] = "SECTOR_ESTADISTICO"
        elif "TIPO DE CULTIVO" in cu or "TIPO CULTIVO" in cu:
            col_map[c] = "TIPO_CULTIVO"
        elif "FENOLOG" in cu:
            col_map[c] = "FENOLOGIA"
        elif "FECHA DE SIEMBRA" in cu or "FECHA SIEMBRA" in cu:
            col_map[c] = "FECHA_SIEMBRA"
        elif "FECHA DE COSECHA" in cu or "FECHA COSECHA" in cu:
            col_map[c] = "FECHA_COSECHA"
        elif "SUPERFICIE SEMBRADA" in cu:
            col_map[c] = "SUP_SEMBRADA"
        elif "SUPERFICIE ASEGURADA" in cu:
            col_map[c] = "SUP_ASEGURADA"
        elif "TIPO DE SINIESTRO" in cu or "TIPO SINIESTRO" in cu:
            col_map[c] = "TIPO_SINIESTRO"
        elif "FECHA DE SINIESTRO" in cu or "FECHA SINIESTRO" in cu:
            col_map[c] = "FECHA_SINIESTRO"
        elif "FECHA DE AVISO" in cu or "FECHA AVISO" in cu:
            col_map[c] = "FECHA_AVISO"
        elif "FECHA DE ATENCIÓN" in cu or "FECHA ATENCION" in cu:
            col_map[c] = "FECHA_ATENCION"
        elif "FECHA" in cu and "PROGRAMACION" in cu and "AJUSTE" in cu and "REPROGRAM" not in cu:
            col_map[c] = "FECHA_PROGRAMACION_AJUSTE"
        elif "FECHA" in cu and "AJUSTE" in cu and "ACTA" in cu and ("ACTA 1" in cu or "ACTA 01" in cu):
            col_map[c] = "FECHA_AJUSTE_ACTA_1"
        elif "FECHA" in cu and "AJUSTE" in cu and "ACTA" in cu and "ACTA 1" not in cu and "ACTA 01" not in cu and "N°" not in cu:
            col_map[c] = "FECHA_AJUSTE_ACTA_FINAL"
        elif "FECHA" in cu and "REPROGRAMACI" in cu and ("01" in cu or cu.endswith("1")) and "02" not in cu and "03" not in cu:
            col_map[c] = "FECHA_REPROGRAMACION_01"
        elif "FECHA" in cu and "REPROGRAMACI" in cu and "02" in cu:
            col_map[c] = "FECHA_REPROGRAMACION_02"
        elif "FECHA" in cu and "REPROGRAMACI" in cu and "03" in cu:
            col_map[c] = "FECHA_REPROGRAMACION_03"
        elif "ESTADO SINIESTRO" in cu:
            col_map[c] = "ESTADO_SINIESTRO"
        elif "ESTADO INSPECCION" in cu or "ESTADO INSPECCIÓN" in cu:
            col_map[c] = "ESTADO_INSPECCION"
        elif "PRIMA NETA" in cu:
            col_map[c] = "PRIMA_NETA_DPTO"
        elif "TIPO DE COBERTURA" in cu or "TIPO COBERTURA" in cu:
            col_map[c] = "TIPO_COBERTURA"
        elif "SUPERFICIE AFECTADA" in cu:
            col_map[c] = "SUP_AFECTADA"
        elif "SUPERFICIE PERDIDA" in cu:
            col_map[c] = "SUP_PERDIDA"
        elif cu == "DICTAMEN" or "DICTAMEN" in cu:
            col_map[c] = "DICTAMEN"
        elif "SUPERFICIE INDEMNIZADA" in cu:
            col_map[c] = "SUP_INDEMNIZADA"
        elif cu == "INDEMNIZACIÓN" or "INDEMNIZACI" in cu:
            col_map[c] = "INDEMNIZACION"
        elif "MONTO DESEMBOLSADO" in cu:
            col_map[c] = "MONTO_DESEMBOLSADO"
        elif "SUPERFICIE DESEMBOLSO" in cu:
            col_map[c] = "SUP_DESEMBOLSO"
        elif "PRODUCTORES" in cu or "N° DE PRODUCTORES" in cu:
            col_map[c] = "N_PRODUCTORES"
        elif "CÓDIGO DE PADRÓN" in cu or "CODIGO DE PADRON" in cu:
            col_map[c] = "CODIGO_PADRON"
        elif "FECHA DE ENVIO" in cu or "FECHA ENVIO" in cu:
            col_map[c] = "FECHA_ENVIO_DRAS"
        elif "FECHA VALIDACI" in cu:
            col_map[c] = "FECHA_VALIDACION"
        elif "FECHA DESEMBOLSO" in cu:
            col_map[c] = "FECHA_DESEMBOLSO"
        elif "PRIORIZADO" in cu:
            col_map[c] = "PRIORIZADO"
    # Evitar duplicados: si dos columnas mapean al mismo nombre, quedarse con la primera
    seen_vals = {}
    dedup_map = {}
    for orig, norm in col_map.items():
        if norm not in seen_vals:
            seen_vals[norm] = orig
            dedup_map[orig] = norm
    df = df.rename(columns=dedup_map)
    # Eliminar columnas duplicadas residuales
    df = df.loc[:, ~df.columns.duplicated()]

    if "DEPARTAMENTO" in df.columns:
        df["DEPARTAMENTO"] = df["DEPARTAMENTO"].astype(str).str.strip().str.upper()
        df = df[~df["DEPARTAMENTO"].isin(["NAN", "", "NONE"])]
    if "TIPO_SINIESTRO" in df.columns:
        df["TIPO_SINIESTRO"] = df["TIPO_SINIESTRO"].astype(str).str.strip().str.upper()

    # Convertir numéricas
    for col in ["SUP_AFECTADA", "SUP_PERDIDA", "SUP_INDEMNIZADA", "INDEMNIZACION",
                 "MONTO_DESEMBOLSADO", "SUP_DESEMBOLSO", "N_PRODUCTORES",
                 "PRIMA_NETA_DPTO", "SUP_SEMBRADA", "SUP_ASEGURADA"]:
        if col in df.columns:
            df[col] = df[col].replace("-", np.nan)
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Coerción de fechas
    _date_cols = ["FECHA_AVISO", "FECHA_ATENCION", "FECHA_SINIESTRO",
                  "FECHA_PROGRAMACION_AJUSTE", "FECHA_AJUSTE_ACTA_1",
                  "FECHA_AJUSTE_ACTA_FINAL", "FECHA_REPROGRAMACION_01",
                  "FECHA_REPROGRAMACION_02", "FECHA_REPROGRAMACION_03",
                  "FECHA_ENVIO_DRAS", "FECHA_VALIDACION", "FECHA_DESEMBOLSO"]
    for col in _date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

    return df


def _normalize_siniestros(uploaded_bytes):
    """Normaliza archivo Sistema de Registro de Siniestros."""
    df = pd.read_excel(uploaded_bytes, header=0, sheet_name=0)

    # Estandarizar columnas
    col_map = {}
    for c in df.columns:
        cu = str(c).strip().upper()
        if "CAMPAÑA" in cu:
            col_map[c] = "CAMPAÑA"
        elif "CODIGO DE AVISO" in cu:
            col_map[c] = "CODIGO_AVISO"
        elif cu == "DEPARTAMENTO":
            col_map[c] = "DEPARTAMENTO"
        elif cu == "PROVINCIA":
            col_map[c] = "PROVINCIA"
        elif cu == "DISTRITO":
            col_map[c] = "DISTRITO"
        elif "SECTOR" in cu:
            col_map[c] = "SECTOR_ESTADISTICO"
        elif "TIPO CULTIVO" in cu:
            col_map[c] = "TIPO_CULTIVO"
        elif "FENOLOG" in cu:
            col_map[c] = "FENOLOGIA"
        elif "FECHA SIEMBRA" in cu:
            col_map[c] = "FECHA_SIEMBRA"
        elif "FECHA COSECHA" in cu:
            col_map[c] = "FECHA_COSECHA"
        elif "SUPERFICIE SEMBRADA" in cu:
            col_map[c] = "SUP_SEMBRADA"
        elif "SUPERFICIE ASEGURADA" in cu:
            col_map[c] = "SUP_ASEGURADA"
        elif "TIPO SINIESTRO" in cu:
            col_map[c] = "TIPO_SINIESTRO"
        elif "FECHA DE SINIESTRO" in cu:
            col_map[c] = "FECHA_SINIESTRO"
        elif "FECHA DE AVISO" in cu:
            col_map[c] = "FECHA_AVISO"
        elif "FECHA DE ATENCIÓN" in cu or "FECHA DE ATENCION" in cu:
            col_map[c] = "FECHA_ATENCION"
        elif "FECHA" in cu and "PROGRAMACION" in cu and "AJUSTE" in cu and "REPROGRAM" not in cu:
            col_map[c] = "FECHA_PROGRAMACION_AJUSTE"
        elif "FECHA" in cu and "AJUSTE" in cu and "ACTA" in cu and ("ACTA 1" in cu or "ACTA 01" in cu):
            col_map[c] = "FECHA_AJUSTE_ACTA_1"
        elif "FECHA" in cu and "AJUSTE" in cu and "ACTA" in cu and "ACTA 1" not in cu and "ACTA 01" not in cu and "N°" not in cu:
            col_map[c] = "FECHA_AJUSTE_ACTA_FINAL"
        elif "FECHA" in cu and "REPROGRAMACI" in cu and ("01" in cu or cu.endswith("1")) and "02" not in cu and "03" not in cu:
            col_map[c] = "FECHA_REPROGRAMACION_01"
        elif "FECHA" in cu and "REPROGRAMACI" in cu and "02" in cu:
            col_map[c] = "FECHA_REPROGRAMACION_02"
        elif "FECHA" in cu and "REPROGRAMACI" in cu and "03" in cu:
            col_map[c] = "FECHA_REPROGRAMACION_03"
        elif "ESTADO SINIESTRO" in cu:
            col_map[c] = "ESTADO_SINIESTRO"
        elif "ESTADO INSPECCION" in cu:
            col_map[c] = "ESTADO_INSPECCION"
        elif "PRIMA NETA" in cu:
            col_map[c] = "PRIMA_NETA_DPTO"
        elif "TIPO COBERTURA" in cu:
            col_map[c] = "TIPO_COBERTURA"
        elif "SUPERFICIE AFECTADA" in cu:
            col_map[c] = "SUP_AFECTADA"
        elif "SUPERFICIE PERDIDA" in cu:
            col_map[c] = "SUP_PERDIDA"
        elif cu == "DICTAMEN":
            col_map[c] = "DICTAMEN"
        elif "SUPERFICIE INDEMNIZADA" in cu:
            col_map[c] = "SUP_INDEMNIZADA"
        elif cu == "INDEMNIZACIÓN" or "INDEMNIZACI" in cu:
            col_map[c] = "INDEMNIZACION"
        elif "MONTO DESEMBOLSADO" in cu:
            col_map[c] = "MONTO_DESEMBOLSADO"
        elif "SUPERFICIE DESEMBOLSO" in cu:
            col_map[c] = "SUP_DESEMBOLSO"
        elif "PRODUCTORES" in cu:
            col_map[c] = "N_PRODUCTORES"
        elif "CÓDIGO DE PADRÓN" in cu or "CODIGO DE PADRON" in cu:
            col_map[c] = "CODIGO_PADRON"
        elif "FECHA ENVIO" in cu:
            col_map[c] = "FECHA_ENVIO_DRAS"
        elif "FECHA VALIDACI" in cu:
            col_map[c] = "FECHA_VALIDACION"
        elif "FECHA DESEMBOLSO" in cu:
            col_map[c] = "FECHA_DESEMBOLSO"
        elif "PRIORIZADO" in cu:
            col_map[c] = "PRIORIZADO"
        elif "OBSERVACI" in cu:
            col_map[c] = "OBSERVACION"
    # Evitar duplicados: si dos columnas mapean al mismo nombre, quedarse con la primera
    seen_vals = {}
    dedup_map = {}
    for orig, norm in col_map.items():
        if norm not in seen_vals:
            seen_vals[norm] = orig
            dedup_map[orig] = norm
    df = df.rename(columns=dedup_map)
    df = df.loc[:, ~df.columns.duplicated()]

    if "DEPARTAMENTO" in df.columns:
        df["DEPARTAMENTO"] = df["DEPARTAMENTO"].astype(str).str.strip().str.upper()
        df = df[~df["DEPARTAMENTO"].isin(["NAN", "", "NONE"])]
    if "TIPO_SINIESTRO" in df.columns:
        df["TIPO_SINIESTRO"] = df["TIPO_SINIESTRO"].astype(str).str.strip().str.upper()

    for col in ["SUP_AFECTADA", "SUP_PERDIDA", "SUP_INDEMNIZADA", "INDEMNIZACION",
                 "MONTO_DESEMBOLSADO", "SUP_DESEMBOLSO", "N_PRODUCTORES",
                 "PRIMA_NETA_DPTO", "SUP_SEMBRADA", "SUP_ASEGURADA"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Coerción de fechas
    _date_cols = ["FECHA_AVISO", "FECHA_ATENCION", "FECHA_SINIESTRO",
                  "FECHA_PROGRAMACION_AJUSTE", "FECHA_AJUSTE_ACTA_1",
                  "FECHA_AJUSTE_ACTA_FINAL", "FECHA_REPROGRAMACION_01",
                  "FECHA_REPROGRAMACION_02", "FECHA_REPROGRAMACION_03",
                  "FECHA_ENVIO_DRAS", "FECHA_VALIDACION", "FECHA_DESEMBOLSO"]
    for col in _date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

    return df


def _normalize_tipo_siniestro_series(series):
    """Normaliza Series de tipo de siniestro (vectorizado, ~50-70% más rápido que .apply)."""
    import unicodedata
    s = series.astype(str).str.strip().str.upper()
    # Reemplazar acentos comunes directamente (más rápido que unicodedata por fila)
    _accent_map = {"Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U", "Ñ": "N",
                   "á": "A", "é": "E", "í": "I", "ó": "O", "ú": "U", "ñ": "N"}
    for accented, plain in _accent_map.items():
        s = s.str.replace(accented, plain, regex=False)
    return s


def process_dynamic_data(midagri_bytes, siniestros_bytes):
    """
    Procesa archivos dinámicos y genera métricas consolidadas.
    MIDAGRI contiene datos de La Positiva (18 dptos).
    Siniestros contiene datos de Rímac (6 dptos).
    Se combinan para obtener el panorama nacional completo.
    """
    midagri = _normalize_midagri(midagri_bytes)
    siniestros = _normalize_siniestros(siniestros_bytes)
    materia = load_materia_asegurada()

    # Normalizar tipo siniestro en ambos (vectorizado)
    if "TIPO_SINIESTRO" in midagri.columns:
        midagri["TIPO_SINIESTRO"] = _normalize_tipo_siniestro_series(midagri["TIPO_SINIESTRO"])
    if "TIPO_SINIESTRO" in siniestros.columns:
        siniestros["TIPO_SINIESTRO"] = _normalize_tipo_siniestro_series(siniestros["TIPO_SINIESTRO"])

    # ═══ COMBINAR ambos datasets en uno solo ═══
    # Guardar referencia de siniestros antes de combinar
    siniestros_solo = siniestros

    # Marcar la empresa de origen ANTES de combinar
    midagri["EMPRESA"] = "LA POSITIVA"
    siniestros["EMPRESA"] = "RIMAC"
    # Usar TODAS las columnas (unión), no solo las comunes
    all_cols = sorted(set(midagri.columns) | set(siniestros.columns))
    combined = pd.concat([
        midagri.reindex(columns=all_cols),
        siniestros.reindex(columns=all_cols)
    ], ignore_index=True)
    # Usar combined como el dataset principal
    midagri = combined
    siniestros = siniestros_solo

    fecha_corte = datetime.now().strftime("%d/%m/%Y")

    # ═══ MÉTRICAS NACIONALES (desde MIDAGRI - todos los avisos) ═══
    total_avisos = len(midagri)

    # Avisos ajustados (estado cerrado/concretado)
    if "ESTADO_INSPECCION" in midagri.columns:
        ajustados = midagri[midagri["ESTADO_INSPECCION"].astype(str).str.upper() == "CERRADO"]
    elif "ESTADO_SINIESTRO" in midagri.columns:
        ajustados = midagri[midagri["ESTADO_SINIESTRO"].astype(str).str.upper() == "CONCRETADO"]
    else:
        ajustados = pd.DataFrame()
    total_ajustados = len(ajustados)
    pct_ajustados = (total_ajustados / total_avisos * 100) if total_avisos > 0 else 0

    # ═══ INDEMNIZACIONES desde MIDAGRI ═══
    ha_indemnizadas = midagri["SUP_INDEMNIZADA"].sum() if "SUP_INDEMNIZADA" in midagri.columns else 0
    monto_indemnizado = midagri["INDEMNIZACION"].sum() if "INDEMNIZACION" in midagri.columns else 0
    monto_desembolsado = midagri["MONTO_DESEMBOLSADO"].sum() if "MONTO_DESEMBOLSADO" in midagri.columns else 0
    # Productores beneficiados: SOLO los de registros con indemnización > 0
    if "N_PRODUCTORES" in midagri.columns and "INDEMNIZACION" in midagri.columns:
        mask_indemn = pd.to_numeric(midagri["INDEMNIZACION"], errors="coerce").fillna(0) > 0
        productores_desembolso = pd.to_numeric(midagri.loc[mask_indemn, "N_PRODUCTORES"], errors="coerce").fillna(0).sum()
    elif "N_PRODUCTORES" in midagri.columns:
        productores_desembolso = midagri["N_PRODUCTORES"].sum()
    else:
        productores_desembolso = 0

    # ═══ DATOS ESTÁTICOS ═══
    prima_total = materia["PRIMA_TOTAL"].sum() if "PRIMA_TOTAL" in materia.columns else 0
    prima_neta = materia["PRIMA_NETA"].sum() if "PRIMA_NETA" in materia.columns else 0
    sup_asegurada = materia["SUPERFICIE_ASEGURADA"].sum() if "SUPERFICIE_ASEGURADA" in materia.columns else 0
    prod_asegurados = materia["PRODUCTORES_ASEGURADOS"].sum() if "PRODUCTORES_ASEGURADOS" in materia.columns else 0

    # Empresas aseguradoras — ahora desde los siniestros consolidados
    if "EMPRESA" in midagri.columns:
        empresas = midagri.groupby("EMPRESA").agg(
            avisos=("EMPRESA", "count"),
            indemnizacion=("INDEMNIZACION", "sum") if "INDEMNIZACION" in midagri.columns else ("EMPRESA", "count"),
            desembolso=("MONTO_DESEMBOLSADO", "sum") if "MONTO_DESEMBOLSADO" in midagri.columns else ("EMPRESA", "count"),
        )
    elif "EMPRESA_ASEGURADORA" in materia.columns:
        empresas = materia.groupby("EMPRESA_ASEGURADORA")["DEPARTAMENTO"].count()
    else:
        empresas = pd.Series()

    # Siniestralidad
    indice_siniestralidad = (monto_indemnizado / prima_neta * 100) if prima_neta > 0 else 0
    pct_desembolso = (monto_desembolsado / monto_indemnizado * 100) if monto_indemnizado > 0 else 0

    # Departamentos con desembolso
    if "MONTO_DESEMBOLSADO" in midagri.columns:
        deptos_con_desembolso = midagri[midagri["MONTO_DESEMBOLSADO"] > 0]["DEPARTAMENTO"].nunique()
    else:
        deptos_con_desembolso = 0

    # ═══ CUADRO 1: Primas y Cobertura por Departamento ═══
    cuadro1 = materia[["DEPARTAMENTO", "PRIMA_TOTAL", "SUPERFICIE_ASEGURADA", "VALORES_ASEGURADOS"]].copy()
    cuadro1 = cuadro1[cuadro1["DEPARTAMENTO"] != "TOTAL"]
    cuadro1 = cuadro1.rename(columns={
        "DEPARTAMENTO": "Departamento",
        "PRIMA_TOTAL": "Prima Total (S/)",
        "SUPERFICIE_ASEGURADA": "Hectáreas Aseguradas",
        "VALORES_ASEGURADOS": "Suma Asegurada Máxima (S/)"
    })
    # Agregar fila TOTAL
    total_row = pd.DataFrame([{
        "Departamento": "TOTAL",
        "Prima Total (S/)": cuadro1["Prima Total (S/)"].sum(),
        "Hectáreas Aseguradas": cuadro1["Hectáreas Aseguradas"].sum(),
        "Suma Asegurada Máxima (S/)": cuadro1["Suma Asegurada Máxima (S/)"].sum()
    }])
    cuadro1 = pd.concat([cuadro1, total_row], ignore_index=True)

    # ═══ CUADRO 2: Indemnizaciones y Desembolsos por Departamento ═══
    if "DEPARTAMENTO" in midagri.columns:
        # Filtrar solo registros con indemnización para contar productores
        midagri_c2 = midagri.copy()
        midagri_c2["_INDEMN_NUM"] = pd.to_numeric(midagri_c2["INDEMNIZACION"], errors="coerce").fillna(0)
        midagri_c2["_PROD_NUM"] = pd.to_numeric(midagri_c2["N_PRODUCTORES"], errors="coerce").fillna(0) if "N_PRODUCTORES" in midagri_c2.columns else 0
        # Solo contar productores donde hay indemnización > 0
        midagri_c2["_PROD_BENEF"] = midagri_c2["_PROD_NUM"].where(midagri_c2["_INDEMN_NUM"] > 0, 0)
        cuadro2 = midagri_c2.groupby("DEPARTAMENTO").agg(
            ha_indemn=("SUP_INDEMNIZADA", "sum"),
            monto_indemn=("INDEMNIZACION", "sum"),
            monto_desemb=("MONTO_DESEMBOLSADO", "sum"),
            productores=("_PROD_BENEF", "sum")
        ).reset_index()
        cuadro2 = cuadro2.rename(columns={
            "DEPARTAMENTO": "Departamento",
            "ha_indemn": "Hectáreas Indemnizadas",
            "monto_indemn": "Monto Indemnizado (S/)",
            "monto_desemb": "Monto Desembolsado (S/)",
            "productores": "Productores con Desembolso"
        })
        cuadro2 = cuadro2.sort_values("Departamento")
        total_row2 = pd.DataFrame([{
            "Departamento": "TOTAL",
            "Hectáreas Indemnizadas": cuadro2["Hectáreas Indemnizadas"].sum(),
            "Monto Indemnizado (S/)": cuadro2["Monto Indemnizado (S/)"].sum(),
            "Monto Desembolsado (S/)": cuadro2["Monto Desembolsado (S/)"].sum(),
            "Productores con Desembolso": cuadro2["Productores con Desembolso"].sum()
        }])
        cuadro2 = pd.concat([cuadro2, total_row2], ignore_index=True)
    else:
        cuadro2 = pd.DataFrame()

    # ═══ CUADRO 3: Eventos de Lluvias Intensas ═══
    if "TIPO_SINIESTRO" in midagri.columns:
        lluvia_df = midagri[midagri["TIPO_SINIESTRO"].isin(LLUVIA_TYPES)]
        total_lluvia = len(lluvia_df)
        pct_lluvia = (total_lluvia / total_avisos * 100) if total_avisos > 0 else 0

        # Conteo por tipo
        lluvia_por_tipo = lluvia_df["TIPO_SINIESTRO"].value_counts()

        # Productores solo de registros con indemnización > 0
        lluvia_c3 = lluvia_df.copy()
        lluvia_c3["_INDEMN_NUM"] = pd.to_numeric(lluvia_c3["INDEMNIZACION"], errors="coerce").fillna(0)
        lluvia_c3["_PROD_NUM"] = pd.to_numeric(lluvia_c3["N_PRODUCTORES"], errors="coerce").fillna(0) if "N_PRODUCTORES" in lluvia_c3.columns else 0
        lluvia_c3["_PROD_BENEF"] = lluvia_c3["_PROD_NUM"].where(lluvia_c3["_INDEMN_NUM"] > 0, 0)
        cuadro3 = lluvia_c3.groupby("DEPARTAMENTO").agg(
            avisos=("CODIGO_AVISO", "count") if "CODIGO_AVISO" in lluvia_c3.columns else ("DEPARTAMENTO", "count"),
            ha_indemn=("SUP_INDEMNIZADA", "sum"),
            monto_indemn=("INDEMNIZACION", "sum"),
            monto_desemb=("MONTO_DESEMBOLSADO", "sum"),
            productores=("_PROD_BENEF", "sum")
        ).reset_index()
        cuadro3 = cuadro3.rename(columns={
            "DEPARTAMENTO": "Departamento",
            "avisos": "Avisos",
            "ha_indemn": "Ha Indemn.",
            "monto_indemn": "Monto Indemnizado (S/)",
            "monto_desemb": "Monto Desembolsado (S/)",
            "productores": "Productores"
        })
        cuadro3 = cuadro3.sort_values("Avisos", ascending=False)
        total_row3 = pd.DataFrame([{
            "Departamento": "TOTAL",
            "Avisos": cuadro3["Avisos"].sum(),
            "Ha Indemn.": cuadro3["Ha Indemn."].sum(),
            "Monto Indemnizado (S/)": cuadro3["Monto Indemnizado (S/)"].sum(),
            "Monto Desembolsado (S/)": cuadro3["Monto Desembolsado (S/)"].sum(),
            "Productores": cuadro3["Productores"].sum()
        }])
        cuadro3 = pd.concat([cuadro3, total_row3], ignore_index=True)
    else:
        cuadro3 = pd.DataFrame()
        total_lluvia = 0
        pct_lluvia = 0
        lluvia_por_tipo = pd.Series()

    # ═══ SINIESTROS POR TIPO (para estadísticas generales) ═══
    if "TIPO_SINIESTRO" in midagri.columns:
        siniestros_por_tipo = midagri["TIPO_SINIESTRO"].value_counts()
        top3_siniestros = siniestros_por_tipo.head(3)
    else:
        siniestros_por_tipo = pd.Series()
        top3_siniestros = pd.Series()

    # ═══ DATOS DEPARTAMENTALES (para ayuda memoria departamental) ═══
    departamentos_list = sorted(midagri["DEPARTAMENTO"].unique().tolist()) if "DEPARTAMENTO" in midagri.columns else []

    return {
        "fecha_corte": fecha_corte,
        "midagri": midagri,
        "siniestros": siniestros,
        "materia": materia,
        # Nacionales
        "total_avisos": int(total_avisos),
        "total_ajustados": int(total_ajustados),
        "pct_ajustados": round(pct_ajustados, 2),
        "ha_indemnizadas": round(ha_indemnizadas, 2),
        "monto_indemnizado": round(monto_indemnizado, 2),
        "monto_desembolsado": round(monto_desembolsado, 2),
        "productores_desembolso": int(productores_desembolso),
        "prima_total": round(prima_total, 2),
        "prima_neta": round(prima_neta, 2),
        "sup_asegurada": round(sup_asegurada, 2),
        "prod_asegurados": int(prod_asegurados),
        "empresas": empresas,
        "indice_siniestralidad": round(indice_siniestralidad, 2),
        "pct_desembolso": round(pct_desembolso, 2),
        "deptos_con_desembolso": deptos_con_desembolso,
        # Cuadros
        "cuadro1": cuadro1,
        "cuadro2": cuadro2,
        "cuadro3": cuadro3,
        # Lluvias
        "total_lluvia": int(total_lluvia),
        "pct_lluvia": round(pct_lluvia, 1),
        "lluvia_por_tipo": lluvia_por_tipo,
        # Siniestros generales
        "siniestros_por_tipo": siniestros_por_tipo,
        "top3_siniestros": top3_siniestros,
        # Departamentos
        "departamentos_list": departamentos_list,
    }


def get_departamento_data(datos, depto):
    """Extrae datos específicos de un departamento para ayuda memoria departamental."""
    depto_upper = depto.strip().upper()
    midagri = datos["midagri"]
    materia = datos["materia"]

    # Filtrar MIDAGRI por departamento (sin copy — solo lectura)
    df_depto = midagri[midagri["DEPARTAMENTO"] == depto_upper]
    mat_depto = materia[materia["DEPARTAMENTO"] == depto_upper]

    # Datos estáticos del departamento
    empresa = mat_depto["EMPRESA_ASEGURADORA"].iloc[0] if len(mat_depto) > 0 and "EMPRESA_ASEGURADORA" in mat_depto.columns else "N/D"
    prima_neta = mat_depto["PRIMA_NETA"].iloc[0] if len(mat_depto) > 0 and "PRIMA_NETA" in mat_depto.columns else 0
    sup_asegurada = mat_depto["SUPERFICIE_ASEGURADA"].iloc[0] if len(mat_depto) > 0 and "SUPERFICIE_ASEGURADA" in mat_depto.columns else 0

    total_avisos = len(df_depto)

    # Indemnizaciones
    ha_indemnizadas = df_depto["SUP_INDEMNIZADA"].sum() if "SUP_INDEMNIZADA" in df_depto.columns else 0
    monto_indemnizado = df_depto["INDEMNIZACION"].sum() if "INDEMNIZACION" in df_depto.columns else 0
    monto_desembolsado = df_depto["MONTO_DESEMBOLSADO"].sum() if "MONTO_DESEMBOLSADO" in df_depto.columns else 0
    productores_desembolso = df_depto["N_PRODUCTORES"].sum() if "N_PRODUCTORES" in df_depto.columns else 0

    # Indemnizables
    if "DICTAMEN" in df_depto.columns:
        indemnizables = len(df_depto[df_depto["DICTAMEN"].astype(str).str.upper() == "INDEMNIZABLE"])
        no_indemnizables = len(df_depto[df_depto["DICTAMEN"].astype(str).str.upper() == "NO INDEMNIZABLE"])
    else:
        indemnizables = 0
        no_indemnizables = 0

    # Avisos por tipo de siniestro
    if "TIPO_SINIESTRO" in df_depto.columns:
        avisos_tipo = df_depto["TIPO_SINIESTRO"].value_counts()
    else:
        avisos_tipo = pd.Series()

    # Distribución por provincia
    if "PROVINCIA" in df_depto.columns:
        dist_provincia = df_depto.groupby("PROVINCIA").agg(
            avisos=("PROVINCIA", "count"),
            sup_indemn=("SUP_INDEMNIZADA", "sum") if "SUP_INDEMNIZADA" in df_depto.columns else ("PROVINCIA", "count"),
            productores=("N_PRODUCTORES", "sum") if "N_PRODUCTORES" in df_depto.columns else ("PROVINCIA", "count"),
            indemniz=("INDEMNIZACION", "sum") if "INDEMNIZACION" in df_depto.columns else ("PROVINCIA", "count"),
            desembolso=("MONTO_DESEMBOLSADO", "sum") if "MONTO_DESEMBOLSADO" in df_depto.columns else ("PROVINCIA", "count"),
        ).reset_index()
        # Calculate % avance (vectorizado, safe contra NaN/inf)
        ind = pd.to_numeric(dist_provincia.get("indemniz", 0), errors="coerce").fillna(0).values
        des = pd.to_numeric(dist_provincia.get("desembolso", 0), errors="coerce").fillna(0).values
        with np.errstate(divide="ignore", invalid="ignore"):
            ratio = np.where(ind > 0, des / ind * 100, 0.0)
        ratio = np.nan_to_num(ratio, nan=0.0, posinf=0.0, neginf=0.0)
        dist_provincia["pct_avance"] = [f"{int(round(p))}%" for p in ratio]
    else:
        dist_provincia = pd.DataFrame()

    # Eventos recientes (último mes)
    eventos_recientes = pd.DataFrame()
    if "FECHA_AVISO" in df_depto.columns or "FECHA_SINIESTRO" in df_depto.columns:
        date_col = "FECHA_AVISO" if "FECHA_AVISO" in df_depto.columns else "FECHA_SINIESTRO"
        try:
            df_depto["_fecha"] = pd.to_datetime(df_depto[date_col], errors="coerce", dayfirst=True)
            now = pd.Timestamp.now()
            last_month = now - pd.Timedelta(days=30)
            recientes = df_depto[df_depto["_fecha"] >= last_month].sort_values("_fecha", ascending=False)
            if len(recientes) > 0:
                cols_evento = []
                for c in ["_fecha", "PROVINCIA", "DISTRITO", "SECTOR_ESTADISTICO", "TIPO_CULTIVO", "TIPO_SINIESTRO", "ESTADO_INSPECCION"]:
                    if c in recientes.columns:
                        cols_evento.append(c)
                eventos_recientes = recientes[cols_evento].head(20)
        except Exception:
            pass

    # Estado de inspección
    if "ESTADO_INSPECCION" in df_depto.columns:
        estados = df_depto["ESTADO_INSPECCION"].value_counts()
    elif "ESTADO_SINIESTRO" in df_depto.columns:
        estados = df_depto["ESTADO_SINIESTRO"].value_counts()
    else:
        estados = pd.Series()

    return {
        "departamento": depto_upper.title(),
        "empresa": empresa,
        "prima_neta": round(float(prima_neta), 2),
        "sup_asegurada": round(float(sup_asegurada), 2),
        "total_avisos": int(total_avisos),
        "ha_indemnizadas": round(float(ha_indemnizadas), 2),
        "monto_indemnizado": round(float(monto_indemnizado), 2),
        "monto_desembolsado": round(float(monto_desembolsado), 2),
        "productores_desembolso": int(productores_desembolso),
        "indemnizables": int(indemnizables),
        "no_indemnizables": int(no_indemnizables),
        "avisos_tipo": avisos_tipo,
        "dist_provincia": dist_provincia,
        "eventos_recientes": eventos_recientes,
        "estados": estados,
        "fecha_corte": datos["fecha_corte"],
        "df_depto": df_depto,
    }


def filter_by_date_range(datos, start_date, end_date):
    """Filtra datos por rango de fechas y recalcula métricas agregadas.
    Retorna nuevo dict datos con midagri filtrado."""
    midagri = datos["midagri"]
    date_col = "FECHA_SINIESTRO" if "FECHA_SINIESTRO" in midagri.columns else "FECHA_AVISO"

    if date_col not in midagri.columns:
        return datos

    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date)
    mask = midagri[date_col].notna() & (midagri[date_col] >= start_ts) & (midagri[date_col] <= end_ts)
    filtered = midagri[mask]

    if len(filtered) == 0:
        return datos

    # Recalcular métricas con el DataFrame filtrado
    new_datos = dict(datos)  # copia superficial
    new_datos["midagri"] = filtered

    new_datos["total_avisos"] = len(filtered)
    new_datos["ha_indemnizadas"] = round(filtered["SUP_INDEMNIZADA"].sum(), 2) if "SUP_INDEMNIZADA" in filtered.columns else 0
    new_datos["monto_indemnizado"] = round(filtered["INDEMNIZACION"].sum(), 2) if "INDEMNIZACION" in filtered.columns else 0
    new_datos["monto_desembolsado"] = round(filtered["MONTO_DESEMBOLSADO"].sum(), 2) if "MONTO_DESEMBOLSADO" in filtered.columns else 0
    new_datos["productores_desembolso"] = int(filtered["N_PRODUCTORES"].sum()) if "N_PRODUCTORES" in filtered.columns else 0

    prima_neta = datos.get("prima_neta", 0)
    indemn = new_datos["monto_indemnizado"]
    new_datos["indice_siniestralidad"] = round(indemn / prima_neta * 100, 2) if prima_neta > 0 else 0
    new_datos["pct_desembolso"] = round(new_datos["monto_desembolsado"] / indemn * 100, 2) if indemn > 0 else 0

    if "TIPO_SINIESTRO" in filtered.columns:
        new_datos["siniestros_por_tipo"] = filtered["TIPO_SINIESTRO"].value_counts()
        new_datos["top3_siniestros"] = new_datos["siniestros_por_tipo"].head(3)

    if "DEPARTAMENTO" in filtered.columns:
        new_datos["departamentos_list"] = sorted(filtered["DEPARTAMENTO"].dropna().unique().tolist())

    return new_datos
